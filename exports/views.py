"""
exports/views.py  —  Rich Excel + PDF export with charts and styled KPI tables.
URL: /api/export/<excel|pdf>/?source=<marketing|channel>&<filters>
"""

import io, datetime, re
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False
    plt = None
from collections import defaultdict

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Max

# ── Colour palette matching dashboard theme ──────────────────────────────────
ORANGE  = '#f58220'
PURPLE  = '#AB1DFE'
TEAL    = '#00BCD4'
GREEN   = '#00C853'
RED     = '#FF3D00'
DARK    = '#1A1235'
MUTED   = '#8a85a5'
YELLOW  = '#FFD600'
COLORS  = [ORANGE, PURPLE, TEAL, GREEN, '#5E35B1', RED, YELLOW, '#E91E8C']

MONTHS = ['Jan','Feb','Mar','Apr','May','Jun',
          'Jul','Aug','Sep','Oct','Nov','Dec']

def _safe(v):
    try:
        f = float(v)
        return 0.0 if f != f else f   # handle NaN
    except (TypeError, ValueError):
        return 0.0

def _fmt(v):
    v = _safe(v)
    if abs(v) >= 1e9: return f"{v/1e9:.2f}B"
    if abs(v) >= 1e6: return f"{v/1e6:.2f}M"
    if abs(v) >= 1e3: return f"{v/1e3:.1f}K"
    return f"{round(v):,}"

def _pct(a, b):
    a, b = _safe(a), _safe(b)
    return round((a/b)*100, 1) if b else 0.0

def _att(a, t):
    return _pct(a, t)

# ─────────────────────────────────────────────────────────────────────────────
#  FILENAME BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _build_filename(source, filters, ext):
    parts = [source.capitalize(), 'KPI_Report']
    for key in ('region','business_unit','arm','franchise','year','month'):
        val = filters.get(key)
        if val:
            clean = re.sub(r'[^\w]', '_', str(val))
            parts.append(clean)
    parts.append(datetime.date.today().strftime('%Y%m%d'))
    return '_'.join(parts) + '.' + ext

# ─────────────────────────────────────────────────────────────────────────────
#  CHART HELPERS  (matplotlib → PNG bytes)
# ─────────────────────────────────────────────────────────────────────────────

def _make_bar_chart(title, labels, datasets, color=ORANGE, width=9, height=3.2,
                    target_data=None, target_label='Target', y_label=''):
    """Return PNG bytes of a bar+line chart."""
    if not HAS_MATPLOTLIB: return None
    fig, ax = plt.subplots(figsize=(width, height))
    fig.patch.set_facecolor('white')
    ax.set_facecolor('#FAFAFA')

    n_sets  = len(datasets)
    n_bars  = len(labels)
    x       = range(n_bars)
    bar_w   = 0.7 / max(n_sets, 1)

    for i, (ds_label, data) in enumerate(datasets):
        offset = (i - (n_sets-1)/2) * bar_w
        xs = [xi + offset for xi in x]
        clr = COLORS[i % len(COLORS)]
        alpha = 1.0 if i == n_sets-1 else 0.45
        bars = ax.bar(xs, data, width=bar_w*0.92, color=clr, alpha=alpha,
                      label=ds_label, zorder=3)
        # Callout labels on latest year only
        if i == n_sets - 1:
            for bar, val in zip(bars, data):
                if val and val > 0:
                    ax.text(bar.get_x() + bar.get_width()/2,
                            bar.get_height() + ax.get_ylim()[1]*0.01,
                            _fmt(val), ha='center', va='bottom',
                            fontsize=7, fontweight='bold', color=DARK)

    # Target line
    if target_data:
        ax.plot(list(x), target_data, color=RED, linewidth=1.8,
                linestyle='--', marker='o', markersize=4,
                label=target_label, zorder=4)

    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, fontsize=8)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: _fmt(v)))
    ax.tick_params(axis='y', labelsize=8)
    ax.set_title(title, fontsize=10, fontweight='bold', color=DARK, pad=8)
    if y_label:
        ax.set_ylabel(y_label, fontsize=8, color=MUTED)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#DDDDDD')
    ax.spines['bottom'].set_color('#DDDDDD')
    ax.yaxis.grid(True, color='#EEEEEE', linewidth=0.7, zorder=0)
    ax.set_axisbelow(True)
    if n_sets > 1 or target_data:
        ax.legend(fontsize=7, loc='upper left', framealpha=0.7)

    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_donut_chart(title, labels, values, colors=None, width=4.5, height=3.5):
    """Return PNG bytes of a donut chart."""
    if not HAS_MATPLOTLIB: return None
    if not any(_safe(v) > 0 for v in values):
        return None
    fig, ax = plt.subplots(figsize=(width, height))
    fig.patch.set_facecolor('white')
    clrs = colors or COLORS[:len(labels)]
    vals = [max(_safe(v), 0) for v in values]
    wedges, _ = ax.pie(vals, colors=clrs, startangle=90,
                       wedgeprops=dict(width=0.55))
    ax.set_title(title, fontsize=9, fontweight='bold', color=DARK, pad=6)
    total = sum(vals)
    legend_labels = [f"{l}: {_fmt(v)} ({_pct(v,total):.1f}%)"
                     for l, v in zip(labels, vals)]
    ax.legend(wedges, legend_labels, loc='lower center',
              bbox_to_anchor=(0.5, -0.22), fontsize=7,
              ncol=1, framealpha=0.7)
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=120, bbox_inches='tight',
                facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf


def _make_gauge_chart(title, value, max_val=100, color=GREEN,
                      width=3.2, height=2.8):
    """Return PNG bytes of a simple gauge/progress arc."""
    if not HAS_MATPLOTLIB: return None
    fig, ax = plt.subplots(figsize=(width, height),
                           subplot_kw=dict(projection='polar'))
    fig.patch.set_facecolor('white')
    pct = min(_safe(value) / max(_safe(max_val), 1), 1.0)
    # Background arc
    theta = [i * 3.14159 / 100 for i in range(101)]
    ax.plot(theta, [1]*101, color='#EEEEEE', linewidth=14, solid_capstyle='round')
    # Value arc
    n = max(int(pct * 100), 1)
    ax.plot(theta[:n], [1]*n, color=color, linewidth=14, solid_capstyle='round')
    ax.set_ylim(0, 1.5)
    ax.set_theta_offset(3.14159)
    ax.set_theta_direction(-1)
    ax.set_thetamin(0); ax.set_thetamax(180)
    ax.axis('off')
    ax.text(0, 0.1, f"{_safe(value):.1f}%", ha='center', va='center',
            fontsize=14, fontweight='bold', color=color,
            transform=ax.transData)
    ax.set_title(title, fontsize=8, fontweight='bold', color=DARK,
                 pad=2, y=1.05)
    plt.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=110, bbox_inches='tight',
                facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
#  DATA COLLECTORS
# ─────────────────────────────────────────────────────────────────────────────

def _get_channel_kpis_and_charts(filters, request=None):
    from channel.models import ChannelDaily
    from channel.views import get_scoped_qs
    # Use RBAC scoping if request available, else full queryset
    if request is not None:
        qs = get_scoped_qs(request.user)
    else:
        qs = ChannelDaily.objects.all()

    r  = filters.get('region')
    bu = filters.get('business_unit')
    ar = filters.get('arm')
    fr = filters.get('franchise')
    yr = filters.get('year')
    mo = filters.get('month')

    if r:  qs = qs.filter(region=r)
    if bu: qs = qs.filter(business_unit=bu)
    if ar: qs = qs.filter(arm=ar)
    if fr: qs = qs.filter(franchise_id=fr)
    if yr: qs = qs.filter(date__year=int(yr))
    if mo: qs = qs.filter(date__month=int(mo))

    # Flow aggregate
    agg = qs.aggregate(
        fca_ach=Sum('fca_ach'), fca_target=Sum('fca_target'),
        g4_ach=Sum('g4_ach'),   g4_target=Sum('g4_target'),
        mnp_ach=Sum('mnp_ach'), mnp_target=Sum('mnp_target'),
        loading_ach=Sum('loading_ach'), loading_target=Sum('loading_target'),
        m0_revenue_ach=Sum('m0_revenue_ach'), m0_revenue_target=Sum('m0_revenue_target'),
        hvc_ach=Sum('hvc_ach'), hvc_target=Sum('hvc_target'),
        bundle_ach=Sum('bundle_ach'), bundle_target=Sum('bundle_target'),
        qos_ach=Sum('qos_ach'), qos_target=Sum('qos_target'),
        sd_bundle=Sum('sd_bundle'), zr=Sum('zr'), zr_fca=Sum('zr_fca'),
        dormancy_count=Sum('dormancy_count'),
        female_fca_count=Sum('female_fca_count'),
        cm_ga=Sum('cm_ga'), uload_recharge_ach=Sum('uload_recharge_ach'),
        cm_disown=Sum('cm_disown'),
        new_sim_sale_disowned_cnics=Sum('new_sim_sale_disowned_cnics'),
        fca_within_90d_disowned=Sum('fca_within_90d_disowned'),
        active_90d_base_disown=Sum('active_90d_base_disown'),
        m0_rev_ga=Sum('m0_rev_ga'), m0_rev_fca=Sum('m0_rev_fca'),
        m0_rev_mnp=Sum('m0_rev_mnp'), m0_rev_mbb=Sum('m0_rev_mbb'),
        m0_rev_data_sim=Sum('m0_rev_data_sim'), m0_hvc_rev=Sum('m0_hvc_rev'),
        fca_m0=Sum('fca_m0'), mbb_ach=Sum('mbb_ach'),
        data_sim_fca=Sum('data_sim_fca'),
        latest_date=Max('date'),
    )

    ld = agg.get('latest_date')
    sq = (qs.filter(date__year=ld.year, date__month=ld.month)
          if ld and not mo else qs)
    stock = sq.aggregate(
        cm_evc_active=Sum('cm_evc_active'),
        cm_evc_active_platinum=Sum('cm_evc_active_platinum'),
        cm_evc_active_gold=Sum('cm_evc_active_gold'),
        cm_evc_active_silver=Sum('cm_evc_active_silver'),
        cm_964_active=Sum('cm_964_active'),
        cm_964_active_platinum=Sum('cm_964_active_platinum'),
        cm_964_active_gold=Sum('cm_964_active_gold'),
        cm_964_active_silver=Sum('cm_964_active_silver'),
        evc_active_base=Sum('evc_active_base'),
        evc_retailer=Sum('evc_retailer'),
        npr=Sum('npr'),
        active_so_daily_avg=Sum('active_so_daily_avg'),
        daily_active_served=Sum('daily_active_served'),
        daily_active_evc=Sum('daily_active_evc'),
    )

    k = {**agg, **stock}
    s = _safe

    def arpu(rev, vol):
        r, v = s(rev), s(vol)
        return round(r/v, 2) if v else 0.0

    fca = s(k.get('fca_ach'))

    # ── KPI sections ─────────────────────────────────────────────────────────
    sections = [
        {
            'title': 'Target Performance',
            'color': ORANGE,
            'rows': [
                ('FCA Achievement',        s(k.get('fca_ach')),        'fca_ach'),
                ('FCA Target',             s(k.get('fca_target')),     None),
                ('FCA Attainment %',       _att(k.get('fca_ach'), k.get('fca_target')), None),
                ('4G Achievement',         s(k.get('g4_ach')),         'g4_ach'),
                ('4G Target',              s(k.get('g4_target')),      None),
                ('4G Attainment %',        _att(k.get('g4_ach'), k.get('g4_target')), None),
                ('4G Penetration % of FCA',_pct(s(k.get('g4_ach')), fca), None),
                ('MNP Achievement',        s(k.get('mnp_ach')),        'mnp_ach'),
                ('MNP Target',             s(k.get('mnp_target')),     None),
                ('MNP Attainment %',       _att(k.get('mnp_ach'), k.get('mnp_target')), None),
                ('Recharge Achievement',   s(k.get('loading_ach')),    'loading_ach'),
                ('Recharge Target',        s(k.get('loading_target')), None),
                ('Recharge Attainment %',  _att(k.get('loading_ach'), k.get('loading_target')), None),
                ('M0 Revenue Achievement', s(k.get('m0_revenue_ach')), 'm0_revenue_ach'),
                ('M0 Revenue Target',      s(k.get('m0_revenue_target')), None),
                ('M0 Revenue Attainment %',_att(k.get('m0_revenue_ach'), k.get('m0_revenue_target')), None),
                ('HVC Achievement',        s(k.get('hvc_ach')),        'hvc_ach'),
                ('HVC Target',             s(k.get('hvc_target')),     None),
                ('HVC Attainment %',       _att(k.get('hvc_ach'), k.get('hvc_target')), None),
                ('Bundle Achievement',     s(k.get('bundle_ach')),     'bundle_ach'),
                ('Bundle Target',          s(k.get('bundle_target')),  None),
                ('Bundle Attainment %',    _att(k.get('bundle_ach'), k.get('bundle_target')), None),
                ('QOS Achievement',        s(k.get('qos_ach')),        None),
                ('QOS Target',             s(k.get('qos_target')),     None),
                ('QOS Attainment %',       _att(k.get('qos_ach'), k.get('qos_target')), None),
            ],
        },
        {
            'title': 'Quality Metrics',
            'color': PURPLE,
            'rows': [
                ('SD Bundle',                   s(k.get('sd_bundle')),          'sd_bundle'),
                ('ZR Count',                    s(k.get('zr')),                 'zr'),
                ('ZR of FCA %',                 _pct(s(k.get('zr_fca')), fca), None),
                ('Dormancy Count',              s(k.get('dormancy_count')),     'dormancy_count'),
                ('Female FCA Count',            s(k.get('female_fca_count')),   'female_fca_count'),
                ('Female FCA % of FCA',         _pct(s(k.get('female_fca_count')), fca), None),
                ('CM Disown',                   s(k.get('cm_disown')),          None),
                ('New SIM Sale Disowned CNICs', s(k.get('new_sim_sale_disowned_cnics')), None),
                ('FCA Within 90D Disowned',     s(k.get('fca_within_90d_disowned')), None),
                ('Active 90D Base Disown',      s(k.get('active_90d_base_disown')), None),
                ('U-Load Recharge Ach',         s(k.get('uload_recharge_ach')), None),
                ('CM GA',                       s(k.get('cm_ga')),              None),
            ],
        },
        {
            'title': 'Enablers (Closing Month)',
            'color': TEAL,
            'rows': [
                ('CM EVC Active',           s(k.get('cm_evc_active')),          'cm_evc_active'),
                ('CM EVC Platinum',         s(k.get('cm_evc_active_platinum')), None),
                ('CM EVC Gold',             s(k.get('cm_evc_active_gold')),     None),
                ('CM EVC Silver',           s(k.get('cm_evc_active_silver')),   None),
                ('CM 964 Active',           s(k.get('cm_964_active')),          'cm_964_active'),
                ('CM 964 Platinum',         s(k.get('cm_964_active_platinum')), None),
                ('CM 964 Gold',             s(k.get('cm_964_active_gold')),     None),
                ('CM 964 Silver',           s(k.get('cm_964_active_silver')),   None),
                ('EVC Active Base',         s(k.get('evc_active_base')),        None),
                ('EVC Retailer',            s(k.get('evc_retailer')),           None),
                ('NPR',                     s(k.get('npr')),                    None),
                ('Active SO Daily Avg',     s(k.get('active_so_daily_avg')),    None),
                ('Daily Active Served',     s(k.get('daily_active_served')),    None),
                ('Daily Active EVC',        s(k.get('daily_active_evc')),       None),
            ],
        },
        {
            'title': 'ARPU & M0 Revenue',
            'color': GREEN,
            'rows': [
                ('ARPU GA',          arpu(k.get('m0_rev_ga'),       k.get('fca_ach')),  None),
                ('ARPU FCA',         arpu(k.get('m0_rev_fca'),      k.get('fca_m0')),   None),
                ('ARPU MNP',         arpu(k.get('m0_rev_mnp'),      k.get('mnp_ach')),  None),
                ('ARPU MBB',         arpu(k.get('m0_rev_mbb'),      k.get('mbb_ach')),  None),
                ('ARPU Data SIM',    arpu(k.get('m0_rev_data_sim'), k.get('data_sim_fca')), None),
                ('ARPU HVC',         arpu(k.get('m0_hvc_rev'),      k.get('hvc_ach')),  None),
                ('M0 Rev GA',        s(k.get('m0_rev_ga')),          None),
                ('M0 Rev FCA',       s(k.get('m0_rev_fca')),         None),
                ('M0 Rev MNP',       s(k.get('m0_rev_mnp')),         None),
                ('M0 Rev MBB',       s(k.get('m0_rev_mbb')),         None),
                ('M0 Rev Data SIM',  s(k.get('m0_rev_data_sim')),    None),
                ('M0 HVC Rev',       s(k.get('m0_hvc_rev')),         None),
                ('FCA M0 Volume',    s(k.get('fca_m0')),             None),
                ('MBB Ach',          s(k.get('mbb_ach')),            None),
                ('Data SIM FCA',     s(k.get('data_sim_fca')),       None),
            ],
        },
    ]

    # ── Monthly trend data for charts ─────────────────────────────────────────
    # Re-use same base queryset (already scoped + filtered), just drop month filter
    if request is not None:
        trend_qs = get_scoped_qs(request.user)
    else:
        trend_qs = ChannelDaily.objects.all()
    if r:  trend_qs = trend_qs.filter(region=r)
    if bu: trend_qs = trend_qs.filter(business_unit=bu)
    if ar: trend_qs = trend_qs.filter(arm=ar)
    if fr: trend_qs = trend_qs.filter(franchise_id=fr)
    if yr: trend_qs = trend_qs.filter(date__year=int(yr))

    chart_metrics = ['fca_ach','fca_target','g4_ach','g4_target',
                     'mnp_ach','mnp_target','loading_ach','loading_target',
                     'm0_revenue_ach','m0_revenue_target',
                     'hvc_ach','hvc_target','bundle_ach','bundle_target',
                     'cm_evc_active','cm_964_active','sd_bundle','zr',
                     'dormancy_count','female_fca_count']

    monthly_rows = (trend_qs.exclude(date__isnull=True)
                    .values('date__year','date__month')
                    .annotate(**{f: Sum(f) for f in chart_metrics}))

    by_year = defaultdict(lambda: {m: [0]*12 for m in chart_metrics})
    for row in monthly_rows:
        yy, mm = row['date__year'], row['date__month']
        for f in chart_metrics:
            by_year[yy][f][mm-1] = _safe(row.get(f))

    years = sorted(by_year.keys())
    # Last valid month index
    def last_idx(metric):
        li = -1
        for yy in years:
            for i, v in enumerate(by_year[yy][metric]):
                if v and v > 0: li = i
        return li if li >= 0 else 11

    def month_labels(li):
        return MONTHS[:li+1]

    def datasets_for(metric, years_list):
        li = last_idx(metric)
        lbs = month_labels(li)
        dsets = [(str(y), by_year[y][metric][:li+1]) for y in years_list]
        return lbs, dsets

    def target_for(metric, tgt_metric, years_list):
        li = last_idx(metric)
        latest_y = years_list[-1] if years_list else None
        if not latest_y: return None
        return by_year[latest_y][tgt_metric][:li+1]

    # ── Build charts ──────────────────────────────────────────────────────────
    charts = []

    chart_defs = [
        ('FCA Monthly Trend',        'fca_ach',        'fca_target',        ORANGE),
        ('4G Achievement Trend',     'g4_ach',         'g4_target',         '#5E35B1'),
        ('MNP Trend',                'mnp_ach',        'mnp_target',        PURPLE),
        ('Recharge Trend',           'loading_ach',    'loading_target',    TEAL),
        ('M0 Revenue Trend',         'm0_revenue_ach', 'm0_revenue_target', GREEN),
        ('HVC Trend',                'hvc_ach',        'hvc_target',        YELLOW),
        ('Bundle Trend',             'bundle_ach',     'bundle_target',     '#E91E8C'),
        ('CM EVC Active Trend',      'cm_evc_active',  None,                TEAL),
        ('CM 964 Active Trend',      'cm_964_active',  None,                PURPLE),
        ('SD Bundle Trend',          'sd_bundle',      None,                ORANGE),
        ('ZR Count Trend',           'zr',             None,                RED),
        ('Dormancy Trend',           'dormancy_count', None,                MUTED),
        ('Female FCA Trend',         'female_fca_count',None,               '#E91E8C'),
    ]

    for title, metric, tgt_metric, color in chart_defs:
        li = last_idx(metric)
        if li < 0: continue
        lbs, dsets = datasets_for(metric, years)
        tgt = target_for(metric, tgt_metric, years) if tgt_metric else None
        buf = _make_bar_chart(title, lbs, dsets, color=color, target_data=tgt,
                              target_label=f'Target {years[-1]}' if tgt else None)
        charts.append((title, buf))

    # EVC tier donut
    evc_vals = [s(k.get('cm_evc_active_platinum')),
                s(k.get('cm_evc_active_gold')),
                s(k.get('cm_evc_active_silver'))]
    if any(v > 0 for v in evc_vals):
        buf = _make_donut_chart('CM EVC Tier Distribution',
                                ['Platinum', 'Gold', 'Silver'], evc_vals,
                                colors=['#FFD700','#C0C0C0','#CD7F32'])
        charts.append(('CM EVC Tier Distribution', buf))

    # Attainment gauges
    gauge_defs = [
        ('FCA Attainment', _att(k.get('fca_ach'), k.get('fca_target')), ORANGE),
        ('4G Attainment',  _att(k.get('g4_ach'),  k.get('g4_target')),  '#5E35B1'),
        ('MNP Attainment', _att(k.get('mnp_ach'), k.get('mnp_target')), PURPLE),
        ('HVC Attainment', _att(k.get('hvc_ach'), k.get('hvc_target')), GREEN),
    ]
    gauge_bufs = []
    for g_title, g_val, g_clr in gauge_defs:
        if g_val > 0:
            color = GREEN if g_val >= 100 else (YELLOW if g_val >= 70 else RED)
            buf = _make_gauge_chart(g_title, g_val, 100, color=color)
            gauge_bufs.append((g_title, buf))

    return sections, charts, gauge_bufs


def _get_marketing_kpis_and_charts(filters, request=None):
    from marketing.models import SiteData

    qs = SiteData.objects.all()
    r  = filters.get('region')
    bu = filters.get('business_unit')
    ar = filters.get('arm')
    fr = filters.get('franchise')
    yr = filters.get('year')
    mo = filters.get('month')

    if r:  qs = qs.filter(region=r)
    if bu: qs = qs.filter(business_unit=bu)
    if ar: qs = qs.filter(arm=ar)
    if fr: qs = qs.filter(key=fr)
    if yr: qs = qs.filter(year=int(yr))
    if mo: qs = qs.filter(month=int(mo))

    flow = qs.aggregate(
        total_revenue=Sum('tot_revn_amt'), total_net_add=Sum('net_add'),
        total_churn=Sum('gross_churn'),    total_revival=Sum('tot_revival'),
        total_fca=Sum('fca'),
        prepaid_digi=Sum('prepaid_dgtl_amount'),
        postpaid_digi=Sum('postpaid_dgtl_amount'),
        total_conv=Sum('conventional_recharge'),
    )
    latest = qs.aggregate(d=Max('year'), m=Max('month'))
    lp = (qs.filter(year=latest['d'], month=latest['m'])
          if latest['d'] else qs.none())
    stock = lp.aggregate(
        total_activations=Sum('act_90d'),
        total_base_4g=Sum('act_90d_4g'),
        total_base_30d=Sum('act_30d'),
        total_hvc=Sum('hvc_base'),
        total_bvs=Sum('bvs_retailer'),
        total_evc=Sum('evc_retailer'),
        total_handset_4g=Sum('handset_4g'),
        total_act_recharge=Sum('act_recharger'),
        total_daily_active=Sum('avg_dly_act'),
        total_rev_lm=Sum('tot_revn_amt'),
    )

    k = {**flow, **stock}
    s = _safe
    digi   = s(k.get('prepaid_digi')) + s(k.get('postpaid_digi'))
    conv   = s(k.get('total_conv'))
    b90    = s(k.get('total_activations'))
    b4g    = s(k.get('total_base_4g'))
    hvc    = s(k.get('total_hvc'))
    b30    = s(k.get('total_base_30d'))
    rev    = s(k.get('total_revenue'))
    rev_lm = s(k.get('total_rev_lm'))
    bvs    = s(k.get('total_bvs'))
    evc    = s(k.get('total_evc'))
    sites  = (qs.exclude(key__isnull=True).exclude(key='')
              .values('key').distinct().count())
    arpu   = (rev_lm / b90) if b90 else 0

    # Tier breakdown
    plat = gold = silv = 0
    try:
        rev_rows = (lp.exclude(key__isnull=True).exclude(key='')
                    .values('key').annotate(sr=Sum('tot_revn_amt')))
        for rr in rev_rows:
            sv = _safe(rr.get('sr'))
            if sv > 800_000:   plat += 1
            elif sv > 500_000: gold += 1
            else:              silv += 1
    except Exception:
        pass

    sections = [
        {
            'title': 'Revenue & Base',
            'color': ORANGE,
            'rows': [
                ('Total Revenue (PKR)',   rev,              None),
                ('Total Net Adds',        s(k.get('total_net_add')), None),
                ('Active 90D Base',       b90,              None),
                ('Active 4G Base',        b4g,              None),
                ('HVC Base',              hvc,              None),
                ('Active 30D Base',       b30,              None),
                ('ARPU (PKR)',            round(arpu,2),    None),
                ('4G Penetration %',      _pct(b4g, b90),   None),
                ('Total Sites',           sites,            None),
            ],
        },
        {
            'title': 'Recharge',
            'color': TEAL,
            'rows': [
                ('Total Recharge (PKR)',       digi + conv,   None),
                ('Digital Recharge (PKR)',     digi,          None),
                ('Conventional Recharge',      conv,          None),
                ('Gross Churn',                s(k.get('total_churn')), None),
                ('Total Revival',              s(k.get('total_revival')), None),
                ('Act Rechargers',             s(k.get('total_act_recharge')), None),
                ('Act Recharger % of 90D',     _pct(s(k.get('total_act_recharge')), b90), None),
                ('Avg Daily Active',           s(k.get('total_daily_active')), None),
                ('Daily Active % of 90D',      _pct(s(k.get('total_daily_active')), b90), None),
            ],
        },
        {
            'title': 'Retailers & Tiers',
            'color': PURPLE,
            'rows': [
                ('BVS Retailers',              bvs,                           None),
                ('EVC Retailers',              evc,                           None),
                ('BVS per Site',               round(bvs/sites,2) if sites else 0, None),
                ('EVC per Site',               round(evc/sites,2) if sites else 0, None),
                ('Handset 4G',                 s(k.get('total_handset_4g')),  None),
                ('Tier Platinum Sites (>800K)',plat,                          None),
                ('Tier Gold Sites (500K-800K)',gold,                          None),
                ('Tier Silver Sites (<500K)',   silv,                          None),
            ],
        },
    ]

    # Monthly trends for charts
    trend_qs = SiteData.objects.all()
    if r:  trend_qs = trend_qs.filter(region=r)
    if bu: trend_qs = trend_qs.filter(business_unit=bu)
    if ar: trend_qs = trend_qs.filter(arm=ar)
    if fr: trend_qs = trend_qs.filter(key=fr)
    if yr: trend_qs = trend_qs.filter(year=int(yr))

    mfields = ['tot_revn_amt','net_add','gross_churn','act_90d',
               'act_90d_4g','hvc_base','prepaid_dgtl_amount',
               'conventional_recharge','act_recharger']

    mrows = (trend_qs.values('year','month')
             .annotate(**{f: Sum(f) for f in mfields}))

    by_yr = defaultdict(lambda: {f: [0]*12 for f in mfields})
    for row in mrows:
        yy, mm = row['year'], row['month']
        for f in mfields:
            by_yr[yy][f][mm-1] = _safe(row.get(f))

    years = sorted(by_yr.keys())

    def last_idx(metric):
        li = -1
        for yy in years:
            for i, v in enumerate(by_yr[yy][metric]):
                if v and v > 0: li = i
        return li if li >= 0 else 11

    def ds_for(metric):
        li = last_idx(metric)
        lbs = MONTHS[:li+1]
        dsets = [(str(y), by_yr[y][metric][:li+1]) for y in years]
        return lbs, dsets

    charts = []
    chart_defs = [
        ('Revenue Monthly Trend',     'tot_revn_amt',            ORANGE),
        ('Net Adds Monthly Trend',    'net_add',                 GREEN),
        ('Gross Churn Trend',         'gross_churn',             RED),
        ('90D Base Trend',            'act_90d',                 TEAL),
        ('4G Base Trend',             'act_90d_4g',              '#5E35B1'),
        ('HVC Base Trend',            'hvc_base',                YELLOW),
        ('Digital Recharge Trend',    'prepaid_dgtl_amount',     PURPLE),
        ('Conventional Recharge',     'conventional_recharge',   '#E91E8C'),
        ('Active Rechargers Trend',   'act_recharger',           TEAL),
    ]
    for title, metric, color in chart_defs:
        li = last_idx(metric)
        if li < 0: continue
        lbs, dsets = ds_for(metric)
        buf = _make_bar_chart(title, lbs, dsets, color=color)
        charts.append((title, buf))

    # Tier donut
    if plat + gold + silv > 0:
        buf = _make_donut_chart('Revenue Tier Distribution',
                                ['Platinum (>800K)', 'Gold (500-800K)', 'Silver (<500K)'],
                                [plat, gold, silv],
                                colors=['#FFD700','#C0C0C0','#CD7F32'])
        charts.append(('Revenue Tier Distribution', buf))

    # 4G penetration gauge
    pen_val = _pct(b4g, b90)
    if pen_val > 0:
        clr = GREEN if pen_val >= 60 else (YELLOW if pen_val >= 40 else RED)
        buf = _make_gauge_chart('4G Penetration', pen_val, 100, color=clr)
        charts.append(('4G Penetration Gauge', buf))

    return sections, charts, []


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _build_excel(sections, charts, gauges, title, filters):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: KPI Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "KPI Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16

    thin = Side(style='thin', color='E8E9ED')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hex_fill(h): return PatternFill("solid", fgColor='FF' + h.lstrip('#').upper())
    def wfont(sz=10, bold=False, color='1A1235'):
        c = ('FF' + color.lstrip('#').upper())[:8]
        return Font(size=sz, bold=bold, color=c, name='Calibri')

    # Title banner
    ws.merge_cells('A1:C1')
    ws['A1'] = title.replace('_', ' ')
    ws['A1'].font      = Font(size=15, bold=True, color='FFFFFFFF', name='Calibri')
    ws['A1'].fill      = hex_fill(DARK)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # Filter info row
    ws.merge_cells('A2:C2')
    filter_str = '  |  '.join(
        f"{k.replace('_',' ').title()}: {v}"
        for k, v in filters.items() if v
    ) or 'All Data'
    ws['A2'] = f"Filters: {filter_str}"
    ws['A2'].font      = Font(size=9, italic=True, color='FF8A85A5', name='Calibri')
    ws['A2'].fill      = hex_fill('#F0F1F5')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    ws['A3'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'].font = Font(size=8, italic=True, color='FF8A85A5')
    ws.row_dimensions[3].height = 14

    current_row = 5
    for sec in sections:
        sec_color = sec['color'].lstrip('#')
        sec_title = sec['title']

        # Section header
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = f"  {sec_title}"
        ws[f'A{current_row}'].font      = Font(size=11, bold=True,
                                                color='FFFFFFFF', name='Calibri')
        ws[f'A{current_row}'].fill      = hex_fill(sec['color'])
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left',
                                                     vertical='center')
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Column sub-headers
        for col, txt in [(1,'KPI'),(2,'Value (Raw)'),(3,'Formatted')]:
            c = ws.cell(row=current_row, column=col, value=txt)
            c.font      = Font(size=9, bold=True, color='FFFFFFFF', name='Calibri')
            c.fill      = hex_fill(ORANGE)
            c.alignment = Alignment(horizontal='center' if col>1 else 'left',
                                    vertical='center')
            c.border    = bdr
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for i, (label, raw, _) in enumerate(sec['rows']):
            bg = 'FFF5EE' if i % 2 == 0 else 'FAFAFA'
            fill = hex_fill(bg)

            lbl_cell = ws.cell(row=current_row, column=1, value=label)
            lbl_cell.font      = wfont(10)
            lbl_cell.fill      = fill
            lbl_cell.border    = bdr
            lbl_cell.alignment = Alignment(horizontal='left', vertical='center',
                                           indent=1)

            raw_cell = ws.cell(row=current_row, column=2, value=raw)
            raw_cell.font      = wfont(10, bold=True, color=sec_color)
            raw_cell.fill      = fill
            raw_cell.border    = bdr
            raw_cell.alignment = Alignment(horizontal='right', vertical='center')
            raw_cell.number_format = '#,##0.00' if isinstance(raw, float) and raw != int(raw) else '#,##0'

            fmt_cell = ws.cell(row=current_row, column=3, value=_fmt(raw))
            fmt_cell.font      = wfont(10)
            fmt_cell.fill      = fill
            fmt_cell.border    = bdr
            fmt_cell.alignment = Alignment(horizontal='right', vertical='center')

            ws.row_dimensions[current_row].height = 17
            current_row += 1

        current_row += 1  # spacer

    ws.freeze_panes = 'A5'

    # ── Sheet 2: Charts ───────────────────────────────────────────────────────
    if charts or gauges:
        ws2 = wb.create_sheet("Charts")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells('A1:L1')
        ws2['A1'] = f"{title.replace('_',' ')} — Monthly Trend Charts"
        ws2['A1'].font      = Font(size=13, bold=True, color='FFFFFFFF', name='Calibri')
        ws2['A1'].fill      = hex_fill(DARK)
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 32

        # Gauges side by side at top
        if gauges:
            ws2.merge_cells('A2:L2')
            ws2['A2'] = "Attainment Gauges"
            ws2['A2'].font = Font(size=10, bold=True, color='FF1A1235')
            ws2.row_dimensions[2].height = 16

            col_offset = 1
            for g_title, g_buf in gauges:
                img = XLImage(g_buf)
                img.width  = 200
                img.height = 175
                col_letter = get_column_letter(col_offset)
                ws2.add_image(img, f'{col_letter}3')
                col_offset += 3

        # Charts in 2-column grid
        CHART_W  = 540   # px
        CHART_H  = 215
        ROW_H_PX = 15
        rows_per_chart = int(CHART_H / ROW_H_PX) + 1
        gauge_rows = 14 if gauges else 0
        start_row  = 3 + gauge_rows

        for ci, (c_title, c_buf) in enumerate(charts):
            row_i = ci // 2
            col_i = ci %  2
            img = XLImage(c_buf)
            img.width  = CHART_W
            img.height = CHART_H
            row_num  = start_row + row_i * rows_per_chart
            col_num  = 1 + col_i * 8
            col_letter = get_column_letter(col_num)
            ws2.add_image(img, f'{col_letter}{row_num}')

        # Column widths for chart sheet
        for ci in range(1, 17):
            ws2.column_dimensions[get_column_letter(ci)].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
#  PDF BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _build_pdf(sections, charts, gauges, title, filters):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, Image as RLImage,
                                     PageBreak, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf    = io.BytesIO()
    PAGE   = landscape(A4)
    L_MAR  = 1.5*cm
    doc    = SimpleDocTemplate(buf, pagesize=PAGE,
                                leftMargin=L_MAR, rightMargin=L_MAR,
                                topMargin=1.2*cm, bottomMargin=1.2*cm)

    styles = getSampleStyleSheet()
    RL_ORANGE = rl_colors.HexColor(ORANGE)
    RL_DARK   = rl_colors.HexColor(DARK)
    RL_MUTED  = rl_colors.HexColor(MUTED)
    RL_WHITE  = rl_colors.white

    title_sty = ParagraphStyle('T', fontName='Helvetica-Bold',
                                fontSize=18, textColor=RL_WHITE,
                                alignment=TA_CENTER, spaceAfter=4)
    sec_sty   = ParagraphStyle('S', fontName='Helvetica-Bold',
                                fontSize=11, textColor=RL_WHITE,
                                alignment=TA_LEFT, leftIndent=6)
    filt_sty  = ParagraphStyle('F', fontName='Helvetica-Oblique',
                                fontSize=8, textColor=RL_MUTED,
                                alignment=TA_CENTER, spaceAfter=6)
    chart_lbl = ParagraphStyle('CL', fontName='Helvetica-Bold',
                                fontSize=9, textColor=RL_DARK,
                                alignment=TA_CENTER, spaceAfter=2)

    PAGE_W = PAGE[0] - 2*L_MAR

    story = []

    # ── Cover / title ─────────────────────────────────────────────────────────
    title_data = [[Paragraph(title.replace('_',' '), title_sty)]]
    title_tbl  = Table(title_data, colWidths=[PAGE_W])
    title_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,-1), RL_DARK),
        ('TOPPADDING',    (0,0),(-1,-1), 12),
        ('BOTTOMPADDING', (0,0),(-1,-1), 12),
        ('ROUNDEDCORNERS', [8]),
    ]))
    story.append(title_tbl)
    story.append(Spacer(1, 0.3*cm))

    filter_str = '  |  '.join(
        f"{k.replace('_',' ').title()}: {v}" for k,v in filters.items() if v
    ) or 'All Data'
    story.append(Paragraph(f"Filters: {filter_str}     Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", filt_sty))
    story.append(HRFlowable(width=PAGE_W, thickness=2, color=RL_ORANGE))
    story.append(Spacer(1, 0.4*cm))

    # ── KPI Tables ────────────────────────────────────────────────────────────
    for sec in sections:
        sec_color = rl_colors.HexColor(sec['color'])

        sec_hdr = [[Paragraph(f"  {sec['title']}", sec_sty)]]
        sec_tbl = Table(sec_hdr, colWidths=[PAGE_W])
        sec_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,-1), sec_color),
            ('TOPPADDING',    (0,0),(-1,-1), 7),
            ('BOTTOMPADDING', (0,0),(-1,-1), 7),
        ]))
        story.append(sec_tbl)
        story.append(Spacer(1, 0.1*cm))

        # Two-column KPI table
        rows = sec['rows']
        half = (len(rows)+1)//2
        left  = rows[:half]
        right = rows[half:]

        col_w = (PAGE_W - 0.3*cm) / 2

        def make_kpi_tbl(rows_list):
            data = [['KPI', 'Value', 'Fmt']]
            for i, (lbl, raw, _) in enumerate(rows_list):
                data.append([lbl, _safe(raw), _fmt(raw)])
            t = Table(data, colWidths=[col_w*0.55, col_w*0.25, col_w*0.20])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0),(-1,0), RL_ORANGE),
                ('TEXTCOLOR',  (0,0),(-1,0), RL_WHITE),
                ('FONTNAME',   (0,0),(-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0,0),(-1,0), 8),
                ('ALIGN',      (0,0),(-1,0), 'CENTER'),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),
                 [rl_colors.HexColor('#FFF5EE'), rl_colors.HexColor('#FAFAFA')]),
                ('FONTNAME', (0,1),(-1,-1), 'Helvetica'),
                ('FONTSIZE', (0,1),(-1,-1), 8),
                ('TEXTCOLOR',(0,1),(-1,-1), RL_DARK),
                ('ALIGN',    (1,1),(-1,-1), 'RIGHT'),
                ('ALIGN',    (0,1),(0,-1),  'LEFT'),
                ('GRID',     (0,0),(-1,-1), 0.4, rl_colors.HexColor('#E8E9ED')),
                ('LEFTPADDING',  (0,0),(-1,-1), 5),
                ('RIGHTPADDING', (0,0),(-1,-1), 5),
                ('TOPPADDING',   (0,0),(-1,-1), 3),
                ('BOTTOMPADDING',(0,0),(-1,-1), 3),
            ]))
            return t

        # Pad right side if unequal
        if len(right) < len(left):
            right = right + [('', '', None)] * (len(left) - len(right))

        pair_data = [[make_kpi_tbl(left), make_kpi_tbl(right)]]
        pair_tbl  = Table(pair_data,
                          colWidths=[col_w, col_w],
                          hAlign='LEFT')
        pair_tbl.setStyle(TableStyle([
            ('VALIGN', (0,0),(-1,-1), 'TOP'),
            ('LEFTPADDING',  (0,0),(-1,-1), 0),
            ('RIGHTPADDING', (0,0),(-1,-1), 3),
        ]))
        story.append(pair_tbl)
        story.append(Spacer(1, 0.35*cm))

    # ── Charts page ───────────────────────────────────────────────────────────
    if charts or gauges:
        story.append(PageBreak())

        chart_title = [[Paragraph("Monthly Trend Charts", title_sty)]]
        ct_tbl = Table(chart_title, colWidths=[PAGE_W])
        ct_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,-1), RL_DARK),
            ('TOPPADDING',    (0,0),(-1,-1), 10),
            ('BOTTOMPADDING', (0,0),(-1,-1), 10),
        ]))
        story.append(ct_tbl)
        story.append(Spacer(1, 0.4*cm))

        # Gauges row
        if gauges:
            story.append(Paragraph("Attainment Gauges", styles['Heading3']))
            gauge_imgs = []
            gw = PAGE_W / max(len(gauges), 1) - 0.3*cm
            gh = gw * 0.88
            for g_title, g_buf in gauges:
                gauge_imgs.append(RLImage(g_buf, width=gw, height=gh))
            g_row  = [gauge_imgs]
            g_tbl  = Table(g_row, colWidths=[gw+0.3*cm]*len(gauges))
            g_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                                        ('ALIGN', (0,0),(-1,-1),'CENTER')]))
            story.append(g_tbl)
            story.append(Spacer(1, 0.5*cm))
            story.append(HRFlowable(width=PAGE_W, thickness=1, color=RL_ORANGE))
            story.append(Spacer(1, 0.3*cm))

        # Charts in 2-per-row grid
        CHART_W_PT = (PAGE_W - 0.4*cm) / 2
        CHART_H_PT = CHART_W_PT * 0.40

        for i in range(0, len(charts), 2):
            row_imgs = []
            for ci in range(i, min(i+2, len(charts))):
                c_title, c_buf = charts[ci]
                img = RLImage(c_buf, width=CHART_W_PT, height=CHART_H_PT)
                row_imgs.append(img)
            if len(row_imgs) == 1:
                row_imgs.append('')

            row_data = [row_imgs]
            row_tbl  = Table(row_data,
                             colWidths=[CHART_W_PT+0.2*cm]*2,
                             hAlign='LEFT')
            row_tbl.setStyle(TableStyle([
                ('VALIGN',       (0,0),(-1,-1),'TOP'),
                ('LEFTPADDING',  (0,0),(-1,-1), 2),
                ('RIGHTPADDING', (0,0),(-1,-1), 2),
                ('BOTTOMPADDING',(0,0),(-1,-1), 6),
            ]))
            story.append(row_tbl)

    doc.build(story)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN VIEW
# ─────────────────────────────────────────────────────────────────────────────

@login_required(login_url='/')
def export_view(request, export_type, source=None):
    """
    Main export view.
    URL: /api/export/<source>/<type>/  e.g. /api/export/channel/excel/
    source is passed via URL path — no query param needed.
    """
    # source from URL path (most reliable), then query param, then Referer
    if not source:
        source = request.GET.get('source', '')
    if not source:
        referer = request.META.get('HTTP_REFERER', '')
        source = 'channel' if '/channel/' in referer else 'marketing'

    filters = {
        'region':        request.GET.get('region', ''),
        'business_unit': request.GET.get('business_unit', ''),
        'arm':           request.GET.get('arm', ''),
        'franchise':     request.GET.get('franchise', ''),
        'year':          request.GET.get('year', ''),
        'month':         request.GET.get('month', ''),
    }

    try:
        if source == 'channel':
            sections, charts, gauges = _get_channel_kpis_and_charts(filters, request)
        else:
            sections, charts, gauges = _get_marketing_kpis_and_charts(filters, request)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        resp = HttpResponse(
            f"Export Error\n{'='*60}\nSource: {source}\n{e}\n\n{tb}",
            content_type='text/plain'
        )
        resp['Content-Disposition'] = 'attachment; filename="export_error.txt"'
        return resp

    title    = _build_title(source, filters)
    filename = _build_filename(source, filters, export_type)

    try:
        if export_type == 'excel':
            data  = _build_excel(sections, charts, gauges, title, filters)
            ct    = ('application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet')
            fname = filename + '.xlsx'
        elif export_type == 'pdf':
            data  = _build_pdf(sections, charts, gauges, title, filters)
            ct    = 'application/pdf'
            fname = filename + '.pdf'
        else:
            return HttpResponse("Unknown export type. Use 'excel' or 'pdf'.", status=400)
    except Exception as e:
        import traceback
        resp = HttpResponse(
            f"Build Error\n{'='*60}\n{e}\n\n{traceback.format_exc()}",
            content_type='text/plain'
        )
        resp['Content-Disposition'] = 'attachment; filename="build_error.txt"'
        return resp

    resp = HttpResponse(data.read(), content_type=ct)
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


@login_required(login_url='/')
def export_view_legacy(request, export_type):
    """Legacy: /api/export/excel/ — infers source from Referer header."""
    referer = request.META.get('HTTP_REFERER', '')
    source  = 'channel' if '/channel/' in referer else 'marketing'
    return export_view(request, export_type, source=source)


def _build_title(source, filters):
    parts = [source.capitalize(), 'KPI Report']
    for k in ('region','business_unit','arm','franchise','year','month'):
        v = filters.get(k)
        if v:
            parts.append(v)
    return ' — '.join(parts)


def _build_filename(source, filters, ext):
    parts = [source.capitalize(), 'KPI_Report']
    for k in ('region','business_unit','arm','franchise','year','month'):
        v = filters.get(k)
        if v:
            parts.append(re.sub(r'[^\w]', '_', str(v)))
    parts.append(datetime.date.today().strftime('%Y%m%d'))
    name = '_'.join(parts)
    return name